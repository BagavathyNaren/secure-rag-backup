# app/server.py

from fastapi import FastAPI, HTTPException, Request, Header, Depends
from fastapi.security import OAuth2PasswordRequestForm
from pydantic import BaseModel, Field
from slowapi import Limiter
from slowapi.util import get_remote_address
from slowapi.errors import RateLimitExceeded
from slowapi.middleware import SlowAPIMiddleware
from slowapi.extension import _rate_limit_exceeded_handler
from fastapi.responses import StreamingResponse
from typing import Optional, List
import json
import traceback
import logging
import os
from datetime import datetime, timezone
from fastapi import UploadFile, File
from langchain_core.documents import Document
import tempfile
import shutil

from langchain_core.runnables import RunnableParallel, RunnablePassthrough
from langchain_core.output_parsers import StrOutputParser

# ============================================================
# DB IMPORTS  (✅ Step 5.4: use Alembic upgrade instead of create_all)
# ============================================================

try:
    from db.seed import seed_users
    from db.migrations import upgrade_head
    DB_AVAILABLE = True
except Exception as e:
    import traceback as tb
    print(f"[STARTUP ERROR] Failed to import database modules: {e}")
    print(tb.format_exc())
    DB_AVAILABLE = False

# ============================================================
# AUTH IMPORTS
# ============================================================

try:
    from app.auth import (
        authenticate_user,
        create_access_token,
        get_current_user,
        require_role,
        EXPIRE_MINS,
        authenticate_user_pg,   
    )
    AUTH_AVAILABLE = True
except Exception as e:
    import traceback as tb
    print(f"[STARTUP ERROR] Failed to import app.auth: {e}")
    print(tb.format_exc())
    AUTH_AVAILABLE = False

    def get_current_user():  # type: ignore
        raise HTTPException(
            status_code=503,
            detail="Authentication module failed to load; JWT auth is unavailable.",
        )

# ============================================================
# SECURE RAG MODULE (production refactor: init happens in startup)
# ============================================================

import app.secure_rag as rag

logger = logging.getLogger(__name__)


def _sanitize_loaded_faiss_index(index_path: str = "faiss_index") -> dict:
    """
    Redacts credential-like patterns inside the already-built FAISS docstore
    and re-saves the index. Logs only counts (never secrets).
    """
    import re

    patterns = [
        re.compile(r"claude-api-key-[a-zA-Z0-9\-]{10,}"),
        re.compile(r"AKIA[0-9A-Z]{10,}"),  
        re.compile(r"(?i)secret.{0,5}access.{0,5}key\s*[:=]\s*\S{10,}")
    ]

    obj = rag._vectorstore
    if obj is None:
        return {"status": "skipped", "reason": "vectorstore_none"}

    # unwrap if it's a retriever wrapper
    vs = getattr(obj, "vectorstore", obj)

    docstore = getattr(vs, "docstore", None)
    if docstore is None or not hasattr(docstore, "_dict"):
        return {"status": "skipped", "reason": "docstore_not_supported"}

    modified_docs = 0
    total_replacements = 0

    for _doc_id, doc in docstore._dict.items():  # type: ignore[attr-defined]
        text = getattr(doc, "page_content", "") or ""
        new_text = text
        rep = 0
        for pat in patterns:
            new_text, c = pat.subn("<REDACTED>", new_text)
            rep += c

        if rep > 0:
            doc.page_content = new_text
            modified_docs += 1
            total_replacements += rep

    saved = False
    if total_replacements > 0:
        vs.save_local(index_path)
        saved = True

    return {
        "status": "ok",
        "modified_docs": modified_docs,
        "total_replacements": total_replacements,
        "saved": saved,
    }

# ============================================================
# 1️⃣  FASTAPI INIT
# ============================================================

app = FastAPI(
    title="Tech Secure RAG API",
    description="Enterprise-secured RAG with JWT auth + RBAC + guardrails",
    version="3.0.0",
)

# ============================================================
# 2️⃣  RATE LIMITING
# ============================================================

limiter = Limiter(key_func=get_remote_address)
app.state.limiter = limiter
app.add_exception_handler(RateLimitExceeded, _rate_limit_exceeded_handler)
app.add_middleware(SlowAPIMiddleware)

# ============================================================
# 3️⃣  REQUEST / RESPONSE MODELS
# ============================================================

class SecureRAGRequest(BaseModel):
    question: str

    class Config:
        json_schema_extra = {
            "example": {"question": "What is the minimum password length?"}
        }

class SecureRAGResponse(BaseModel):
    answer: str
    confidence: str
    cached: bool = False

class LoginResponse(BaseModel):
    access_token: str
    token_type: str
    expires_in_minutes: int
    user_id: str
    email: str
    role: str

# ============================================================
# 3b️⃣  EVALUATION MODELS
# ============================================================

class EvalCase(BaseModel):
    id: Optional[str] = None
    question: str
    role: str = "employee"
    should_block: bool = False
    expect_confidence: Optional[str] = Field(default=None)
    expect_answer_contains: List[str] = Field(default_factory=list)
    expect_sources_contains: List[str] = Field(default_factory=list)

class EvalRequest(BaseModel):
    cases: List[EvalCase]
    include_answer: bool = Field(default=False)
    include_sources: bool = Field(default=True)

# ============================================================
# EVAL AUTH GUARD
# ============================================================

EVAL_API_KEY = os.getenv("EVAL_API_KEY", "")

def _require_eval_key(x_eval_key: Optional[str]):
    if not EVAL_API_KEY:
        raise HTTPException(
            status_code=403,
            detail="Evaluation endpoint is disabled. Set EVAL_API_KEY secret.",
        )
    if not x_eval_key or x_eval_key != EVAL_API_KEY:
        raise HTTPException(
            status_code=403,
            detail="Forbidden: invalid or missing x-eval-key header.",
        )

# ============================================================
# Pydantic Model to ingest into Neon pgvector at runtime.
# ============================================================
class IndexDoc(BaseModel):
    text: str = Field(..., min_length=1)
    metadata: dict = Field(default_factory=dict)

class IndexRequest(BaseModel):
    docs: List[IndexDoc]

# ============================================================
# 4️⃣  STARTUP EVENT (single unified startup log sequence)
# ============================================================

@app.on_event("startup")
async def startup():
    # 1) PostgreSQL migrations + seed
    if DB_AVAILABLE:
        try:
            import anyio
            await anyio.to_thread.run_sync(upgrade_head)
            rag.audit.log("DB_MIGRATIONS", "startup", {"status": "upgraded_to_head"})
            from db.url_utils import validate_database_url, redact_database_url
            raw = os.getenv("DATABASE_URL", "").strip()
            validate_database_url(raw)

            # Strip query params before logging (they may contain internal details)
            safe_url = redact_database_url(raw)
            if "?" in safe_url:
                 safe_url = safe_url.split("?")[0]

            rag.audit.log("DATABASE_URL_OK", "startup", {"database_url": safe_url})

        except Exception as e:
            rag.audit.log("DB_MIGRATIONS_ERROR", "startup", {"error": str(e)})
            raise  # fail-fast recommended for production

        try:
            await seed_users()
            rag.audit.log("DB_SEED", "startup", {"status": "complete"})
        except Exception as e:
            rag.audit.log("DB_SEED_ERROR", "startup", {"error": str(e)})
    else:
        rag.audit.log("DB_INIT_SKIPPED", "startup", {"reason": "db module failed to import"})

    # 2) Cache init (Redis if available else memory)
    rag.init_cache()
    prev = rag.read_last_shutdown_marker()
    rag.audit.log("PREVIOUS_SHUTDOWN_MARKER", "startup", {"last_shutdown_ts": prev})

    # 3) Vectorstore init (FAISS or PGVector)
    backend = os.getenv("VECTORSTORE_BACKEND", "faiss").lower()

    # Only meaningful for FAISS
    force_rebuild = (backend == "faiss") and (os.getenv("REBUILD_FAISS", "0").strip() == "1")

    rag.init_vectorstore(force_rebuild=force_rebuild)

    # Only sanitize FAISS indexes (PGVector has no FAISS docstore)
    if backend == "faiss":
          index_path = os.getenv("FAISS_INDEX_PATH", "faiss_index")
          result = _sanitize_loaded_faiss_index(index_path)
          rag.audit.log("FAISS_SANITIZE", "startup", result)

    # 4) LLM init
    rag.init_llm()

    rag.audit.log("STARTUP_COMPLETE", "startup", {"status": "ready"})


def _require_rag_ready():
    # Debug: force "not ready" to test 503 behavior
    if os.getenv("FORCE_RAG_NOT_READY", "0").strip() == "1":
        raise HTTPException(
            status_code=503,
            detail="Service initializing. Please retry in a few seconds.",
        )

    if rag.llm_cache is None or rag._vectorstore is None or rag._embeddings is None or rag._llm is None:
        raise HTTPException(
            status_code=503,
            detail="Service initializing. Please retry in a few seconds.",
        )

@app.on_event("shutdown")
async def shutdown():
    rag.audit.log("SHUTDOWN_START", "system", {})

    # Write marker FIRST (so next startup can prove shutdown happened)
    try:
        ts = datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")
        rag.write_last_shutdown_marker(ts)
        rag.audit.log("SHUTDOWN_MARKER_WRITTEN", "system", {"ts": ts})
    except Exception as e:
        rag.audit.log("SHUTDOWN_MARKER_ERROR", "system", {"error": str(e)})

    # Close RAG resources (Redis connections, etc.)
    try:
        rag.shutdown_rag()
        rag.audit.log("RAG_SHUTDOWN_OK", "system", {})
    except Exception as e:
        rag.audit.log("RAG_SHUTDOWN_ERROR", "system", {"error": str(e)})

    # Dispose DB engine (if present)
    try:
        from db.connection import engine
        await engine.dispose()
        rag.audit.log("DB_ENGINE_DISPOSED", "system", {})
    except Exception as e:
        rag.audit.log("DB_ENGINE_DISPOSE_ERROR", "system", {"error": str(e)})

    rag.audit.log("SHUTDOWN_COMPLETE", "system", {})

# ============================================================
# 5️⃣  AUTH ENDPOINTS
# ============================================================

if AUTH_AVAILABLE:

    @app.post("/auth/login", response_model=LoginResponse, tags=["Authentication"])
    @limiter.limit("10/minute")
    async def login(
        request: Request,
        form_data: OAuth2PasswordRequestForm = Depends(),
    ):
        """
        Login with username + password.
        Returns JWT token valid for EXPIRE_MINS minutes.
        Use token as: Authorization: Bearer <token>
        """
        _require_rag_ready()  # ✅ Step 4

        user = await authenticate_user_pg(form_data.username, form_data.password)
        if not user:
            rag.audit.log("LOGIN_FAILED", "auth", {
                "username": form_data.username,
                "reason": "invalid_credentials",
            })
            raise HTTPException(
                status_code=401,
                detail="Incorrect username or password.",
                headers={"WWW-Authenticate": "Bearer"},
            )

        token = create_access_token(user)

        rag.audit.log("LOGIN_SUCCESS", "auth", {
            "user_id": user["user_id"],
            "username": user["username"],
            "email": user["email"],
            "role": user["role"],
        })

        return LoginResponse(
            access_token=token,
            token_type="bearer",
            expires_in_minutes=EXPIRE_MINS,
            user_id=user["user_id"],
            email=user["email"],
            role=user["role"],
        )

    @app.get("/auth/me", tags=["Authentication"])
    async def whoami(current_user: dict = Depends(get_current_user)):
        return {
            "user_id": current_user["user_id"],
            "email": current_user["email"],
            "role": current_user["role"],
            "sub": current_user["sub"],
        }

else:

    @app.get("/auth/status", tags=["Authentication"])
    async def auth_status():
        return {"status": "AUTH_MODULE_FAILED_TO_LOAD", "detail": "Check container logs for the real error."}

# ============================================================
# 6️⃣  SECURE RAG ENDPOINT (JWT PROTECTED + STREAMING)
# ============================================================

@app.post(
    "/secure-rag/invoke",
    tags=["RAG"],
    summary="Ask a question (JWT required)",
    description=(
        "Submit a question. Role is determined automatically from your JWT token.\n\n"
        "**How to authenticate:**\n"
        "1. POST to `/auth/login` with your username + password\n"
        "2. Copy the `access_token` from the response\n"
        "3. Add header: `Authorization: Bearer <token>`"
    ),
)
@limiter.limit("10/minute")
async def secure_rag_endpoint(
    request: Request,
    body: SecureRAGRequest,
    current_user: dict = Depends(get_current_user),
):
    _require_rag_ready()

    trace_id = rag.new_trace_id()

    user_role = current_user["role"]
    user_id = current_user["user_id"]
    user_email = current_user["email"]

    rag.audit.log("REQUEST_START", trace_id, {
        "user_id": user_id,
        "email": user_email,
        "role": user_role,
        "question_preview": body.question[:120],
    })

    try:
        rag.detect_prompt_injection(body.question, trace_id)
        user_input = rag.redact_pii(body.question, trace_id)

        # Cache lookup (HIGH confidence only)
        cached = rag.llm_cache.get(user_role, user_input)  # type: ignore[union-attr]
        if cached:
            cached_conf = (cached.get("confidence") or "").upper()
            if cached_conf == "HIGH":
                rag.audit.log("CACHE_HIT", trace_id, {
                    "user_id": user_id,
                    "role": user_role,
                    "confidence": cached.get("confidence"),
                })

                async def stream_cached():
                    for word in cached["answer"].split(" "):
                        token_payload = {"token": word + " "}
                        yield f"data: {json.dumps(token_payload)}\n\n"
                    done_payload = {
                        "done": True,
                        "answer": cached["answer"],
                        "confidence": cached["confidence"],
                        "cached": True,
                    }
                    yield f"data: {json.dumps(done_payload)}\n\n"

                return StreamingResponse(stream_cached(), media_type="text/event-stream")

            rag.audit.log("CACHE_BYPASS", trace_id, {
                "user_id": user_id,
                "role": user_role,
                "confidence": cached.get("confidence"),
                "reason": "cached_confidence_not_high",
            })
            rag.llm_cache.invalidate(user_role, user_input)  # type: ignore[union-attr]

        rag.audit.log("CACHE_MISS", trace_id, {"user_id": user_id, "role": user_role})

        # Retrieval
        retriever = rag.build_secure_retriever(user_role, trace_id)
        retrieved_docs = retriever(user_input)
        context = "\n\n".join(doc.page_content for doc in retrieved_docs)

        retrieved_sources = [d.metadata.get("file_name") for d in retrieved_docs]
        sources_str = ", ".join(sorted(set(s for s in retrieved_sources if s))) or "none"

        setup = RunnableParallel(
            context=lambda _: context,
            question=RunnablePassthrough(),
            role=lambda _: user_role,
            sources=lambda _: sources_str,
        )

        rag_chain = setup | rag.secure_prompt | rag._llm | StrOutputParser()

        async def stream_tokens():
            full_answer = ""

            # Block before generation if context contains credentials
            try:
                rag.scan_context_for_credentials(context, trace_id)
            except ValueError as ve:
                rag.audit.log("REQUEST_FAILED", trace_id, {
                    "user_id": user_id,
                    "error": str(ve),
                    "stage": "context_scan",
                })
                yield f"data: {json.dumps({'error': str(ve)})}\n\n"
                return

            async for chunk in rag_chain.astream(user_input):
                full_answer += chunk
                yield f"data: {json.dumps({'token': chunk})}\n\n"

            # Output guardrails + confidence + caching policy
            try:
                full_answer = rag.pre_filter_check(full_answer, trace_id)
                full_answer = rag.scan_answer_for_sensitive_terms(full_answer, trace_id)
                full_answer = rag.model_guard_check(full_answer, context, trace_id)

                confidence = rag.compute_confidence(retrieved_docs, full_answer)

                if confidence == "HIGH":
                    rag.llm_cache.set(user_role, user_input, {  # type: ignore[union-attr]
                        "answer": full_answer,
                        "confidence": confidence,
                    })
                    rag.audit.log("CACHE_STORED", trace_id, {
                        "user_id": user_id,
                        "role": user_role,
                        "confidence": confidence,
                    })
                else:
                    rag.audit.log("CACHE_SKIPPED", trace_id, {
                        "user_id": user_id,
                        "role": user_role,
                        "confidence": confidence,
                        "reason": "low_confidence",
                    })

                rag.audit.log("ANSWER", trace_id, {
                    "user_id": user_id,
                    "email": user_email,
                    "role": user_role,
                    "message": full_answer[:500],
                    "confidence": confidence,
                    "cached": False,
                    "sources": retrieved_sources,
                })

                done_payload = {
                    "done": True,
                    "answer": full_answer,
                    "confidence": confidence,
                    "cached": False,
                }
                yield f"data: {json.dumps(done_payload)}\n\n"

            except ValueError as ve:
                rag.audit.log("REQUEST_FAILED", trace_id, {
                    "user_id": user_id,
                    "error": str(ve),
                    "stage": "output_guard",
                })
                yield f"data: {json.dumps({'error': str(ve)})}\n\n"

        return StreamingResponse(stream_tokens(), media_type="text/event-stream")

    except ValueError as ve:
        rag.audit.log("REQUEST_FAILED", trace_id, {
            "user_id": user_id,
            "error": str(ve),
            "stage": "input_guard",
        })
        raise HTTPException(status_code=400, detail=str(ve))

    except Exception as e:
        rag.audit.log("REQUEST_FAILED", trace_id, {
            "user_id": user_id,
            "error": str(e),
            "stage": "unhandled",
        })
        logger.error(f"Unhandled error: {e}\n{traceback.format_exc()}")
        raise HTTPException(status_code=500, detail="Internal Server Error")

@app.post("/internal/index", tags=["Internal"])
@limiter.limit("2/minute")
async def internal_index(
    request: Request,
    body: IndexRequest,
    x_eval_key: Optional[str] = Header(default=None, alias="x-eval-key"),
):
    _require_rag_ready()
    _require_eval_key(x_eval_key)

    trace_id = rag.new_trace_id()

    # Build LangChain Documents
    from langchain_core.documents import Document

    docs = [
        Document(page_content=d.text, metadata=(d.metadata or {}))
        for d in body.docs
    ]

    stored = rag.index_documents(docs, trace_id=trace_id)

    return {
        "status": "ok",
        "stored": stored,
        "trace_id": trace_id,
        "backend": os.getenv("VECTORSTORE_BACKEND", "faiss").lower(),
    }

@app.post("/internal/upload", tags=["Internal"])
@limiter.limit("2/minute")
async def internal_upload(
    request: Request,
    file: UploadFile = File(...),
    x_eval_key: Optional[str] = Header(default=None, alias="x-eval-key"),
):
    """
    Upload a file (.txt, .pdf, .docx, .csv, .xlsx), parse it, chunk it, and index into PGVector.
    """
    _require_rag_ready()
    _require_eval_key(x_eval_key)

    trace_id = rag.new_trace_id()
    filename = file.filename or "uploaded_file"

    # Save uploaded file to a temp path
    with tempfile.NamedTemporaryFile(delete=False, suffix=os.path.splitext(filename)[1]) as tmp:
        shutil.copyfileobj(file.file, tmp)
        tmp_path = tmp.name

    try:
        # Parse based on extension
        ext = filename.lower().split(".")[-1]
        docs: List[Document] = []

        if ext == "txt":
            from langchain_community.document_loaders import TextLoader
            loader = TextLoader(tmp_path)
            docs = loader.load()

        elif ext == "pdf":
            from app.ingestion import load_pdf
            docs = load_pdf(tmp_path)

        elif ext == "docx":
            from langchain_community.document_loaders import Docx2txtLoader
            loader = Docx2txtLoader(tmp_path)
            docs = loader.load()

        elif ext == "csv":
            from langchain_community.document_loaders import CSVLoader
            loader = CSVLoader(tmp_path)
            docs = loader.load()

        elif ext == "xlsx":
            import openpyxl
            wb = openpyxl.load_workbook(tmp_path)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                headers = [cell.value for cell in ws[1]]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if all(v is None for v in row):
                        continue
                    row_data = dict(zip(headers, row))
                    page_content = ", ".join(f"{k}: {v}" for k, v in row_data.items() if k)
                    docs.append(Document(
                        page_content=page_content,
                        metadata={"file_name": filename, "sheet_name": sheet_name}
                    ))
        else:
            raise HTTPException(status_code=400, detail=f"Unsupported file type: {ext}")

        # Add file_name to metadata
        for doc in docs:
            doc.metadata["file_name"] = filename

        # Chunk then index
        from app.chunking import recursive_character_chunking
        chunks = recursive_character_chunking(docs, chunk_size=600, chunk_overlap=150)

        stored = rag.index_documents(chunks, trace_id=trace_id)

        return {
            "status": "ok",
            "filename": filename,
            "parsed_docs": len(docs),
            "chunks": len(chunks),
            "stored": stored,
            "trace_id": trace_id,
        }

    finally:
        os.unlink(tmp_path)

# ============================================================
# 7️⃣  EVALUATION ENDPOINT
# ============================================================

@app.post("/secure-rag/eval", tags=["Internal"])
@limiter.limit("5/minute")
async def secure_rag_eval(
    request: Request,
    body: EvalRequest,
    x_eval_key: Optional[str] = Header(default=None, alias="x-eval-key"),
):
    _require_rag_ready()
    _require_eval_key(x_eval_key)

    batch_id = rag.new_trace_id()
    rag.audit.log("EVAL_BATCH_START", batch_id, {"total_cases": len(body.cases)})

    results = []
    passed = 0

    for i, case in enumerate(body.cases):
        case_trace_id = rag.new_trace_id()
        case_id = case.id or f"case_{i + 1}"

        rag.audit.log("EVAL_CASE_START", case_trace_id, {
            "batch_id": batch_id,
            "case_id": case_id,
            "role": case.role,
            "should_block": case.should_block,
            "question_preview": case.question[:120],
        })

        answer = None
        confidence = None
        retrieved_sources = []
        blocked = False
        error_msg = None

        try:
            rag.detect_prompt_injection(case.question, case_trace_id)
            clean_input = rag.redact_pii(case.question, case_trace_id)

            retriever = rag.build_secure_retriever(case.role, case_trace_id)
            retrieved_docs = retriever(clean_input)
            retrieved_sources = [d.metadata.get("file_name") for d in retrieved_docs]
            context = "\n\n".join(d.page_content for d in retrieved_docs)

            rag.scan_context_for_credentials(context, case_trace_id)

            sources_str = ", ".join(sorted(set(s for s in retrieved_sources if s))) or "none"

            setup = RunnableParallel(
                context=lambda _: context,
                question=RunnablePassthrough(),
                role=lambda _: case.role,
                sources=lambda _: sources_str,
            )
            rag_chain = setup | rag.secure_prompt | rag._llm | StrOutputParser()
            answer = await rag_chain.ainvoke(clean_input)

            answer = rag.pre_filter_check(answer, case_trace_id)
            answer = rag.scan_answer_for_sensitive_terms(answer, case_trace_id)
            answer = rag.model_guard_check(answer, context, case_trace_id)
            confidence = rag.compute_confidence(retrieved_docs, answer)

        except Exception as e:
            blocked = True
            error_msg = str(e)
            confidence = "BLOCKED"

        checks = {}
        checks["block_check"] = (blocked == case.should_block)
        checks["confidence_check"] = (confidence == case.expect_confidence) if case.expect_confidence else True

        if case.expect_answer_contains and answer:
            checks["answer_content_check"] = all(kw.lower() in answer.lower() for kw in case.expect_answer_contains)
        elif case.expect_answer_contains and not answer:
            checks["answer_content_check"] = False
        else:
            checks["answer_content_check"] = True

        checks["sources_check"] = (
            all(s in set(retrieved_sources) for s in case.expect_sources_contains)
            if case.expect_sources_contains else True
        )

        case_passed = all(checks.values())
        if case_passed:
            passed += 1

        rag.audit.log("EVAL_CASE_RESULT", case_trace_id, {
            "batch_id": batch_id,
            "case_id": case_id,
            "passed": case_passed,
            "blocked": blocked,
            "confidence": confidence,
            "checks": checks,
            "sources": retrieved_sources,
            "error": error_msg,
        })

        result = {
            "case_id": case_id,
            "trace_id": case_trace_id,
            "role": case.role,
            "question": case.question,
            "passed": case_passed,
            "blocked": blocked,
            "confidence": confidence,
            "checks": checks,
            "error": error_msg,
        }
        if body.include_sources:
            result["sources"] = retrieved_sources
        if body.include_answer:
            result["answer"] = answer

        results.append(result)

    total = len(body.cases)
    pass_rate = round((passed / total) * 100, 1) if total > 0 else 0.0

    summary = {
        "batch_id": batch_id,
        "total": total,
        "passed": passed,
        "failed": total - passed,
        "pass_rate_pct": pass_rate,
        "results": results,
    }

    rag.audit.log("EVAL_BATCH_DONE", batch_id, {
        "total": total,
        "passed": passed,
        "failed": total - passed,
        "pass_rate_pct": pass_rate,
    })
    return summary

# ============================================================
# 8️⃣  CACHE ENDPOINTS
# ============================================================

@app.get("/secure-rag/cache/stats", tags=["Internal"])
async def cache_stats(
    x_eval_key: Optional[str] = Header(default=None, alias="x-eval-key"),
):
    _require_rag_ready()
    _require_eval_key(x_eval_key)
    stats = rag.llm_cache.stats()  # type: ignore[union-attr]
    rag.audit.log("CACHE_STATS_REQUESTED", "system", stats)
    return stats

@app.delete("/secure-rag/cache/clear", tags=["Internal"])
async def cache_clear(
    x_eval_key: Optional[str] = Header(default=None, alias="x-eval-key"),
):
    _require_rag_ready()
    _require_eval_key(x_eval_key)
    cleared = rag.llm_cache.clear()  # type: ignore[union-attr]
    rag.audit.log("CACHE_CLEARED", "system", {"entries_removed": cleared})
    return {"status": "cleared", "entries_removed": cleared}

# ============================================================
# 8b️⃣  LIFECYCLE / SHUTDOWN MARKER ENDPOINTS (Internal)
# ============================================================

@app.get("/internal/lifecycle/shutdown-marker", tags=["Internal"])
async def lifecycle_read_shutdown_marker(
    x_eval_key: Optional[str] = Header(default=None, alias="x-eval-key"),
):
    _require_rag_ready()
    _require_eval_key(x_eval_key)

    val = rag.read_last_shutdown_marker()
    backend = None
    try:
        backend = rag.llm_cache.stats().get("backend") if rag.llm_cache else None
    except Exception:
        backend = None

    return {
        "last_shutdown_ts": val,
        "cache_backend": backend,
    }

@app.post("/internal/lifecycle/shutdown-marker", tags=["Internal"])
async def lifecycle_write_shutdown_marker(
    x_eval_key: Optional[str] = Header(default=None, alias="x-eval-key"),
):
    _require_rag_ready()
    _require_eval_key(x_eval_key)

    ts = datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")
    rag.write_last_shutdown_marker(ts)
    read_back = rag.read_last_shutdown_marker()

    return {
        "written_ts": ts,
        "read_back": read_back,
    }

@app.post("/internal/lifecycle/terminate", tags=["Internal"])
async def lifecycle_terminate(
    x_eval_key: Optional[str] = Header(default=None, alias="x-eval-key"),
):
    _require_rag_ready()
    _require_eval_key(x_eval_key)

    import signal
    import threading
    import time

    pid = os.getpid()

    def _killer():
        time.sleep(0.3)
        os.kill(pid, signal.SIGTERM)

    threading.Thread(target=_killer, daemon=True).start()

    return {
        "status": "terminating",
        "pid": pid,
        "note": "Server will receive SIGTERM; shutdown event should run if platform allows graceful stop.",
    }

# ============================================================
# 9️⃣  HEALTH CHECKS
# ============================================================

@app.get("/", tags=["Health"])
def health():
    return {
        "status": "ok",
        "service": "Tech Secure RAG",
        "version": "3.0.0",
        "auth_loaded": AUTH_AVAILABLE,
        "db_loaded": DB_AVAILABLE,
        "rag_ready": (rag.llm_cache is not None and rag._vectorstore is not None and rag._llm is not None),
        "endpoint": "/secure-rag/invoke",
    }

@app.get("/health", tags=["Health"])
async def health_check():
    health_status = {
        "status": "ok",
        "service": "Tech Secure RAG",
        "version": "3.0.0",
        "components": {},
    }

    # PostgreSQL
    if DB_AVAILABLE:
        try:
            from db.connection import engine
            import sqlalchemy
            async with engine.connect() as conn:
                await conn.execute(sqlalchemy.text("SELECT 1"))
            health_status["components"]["postgresql"] = "healthy"
        except Exception as e:
            health_status["components"]["postgresql"] = f"error: {str(e)}"
            health_status["status"] = "degraded"
    else:
        health_status["components"]["postgresql"] = "not_loaded"
        health_status["status"] = "degraded"

    # Vectorstore (FAISS or PGVector)
    try:
        backend = os.getenv("VECTORSTORE_BACKEND", "faiss").lower()
        if rag._vectorstore is not None:
               _ = rag._vectorstore.similarity_search("test", k=1)
               health_status["components"]["vectorstore"] = f"healthy ({backend})"
        else:
               health_status["components"]["vectorstore"] = "not_loaded"
               health_status["status"] = "degraded"
    except Exception as e:
        health_status["components"]["vectorstore"] = f"error: {str(e)}"
        health_status["status"] = "degraded"

    # Embeddings
    try:
        if rag._embeddings is not None:
            rag._embeddings.embed_query("health check")
            health_status["components"]["embeddings"] = "healthy"
        else:
            health_status["components"]["embeddings"] = "not_loaded"
            health_status["status"] = "degraded"
    except Exception as e:
        health_status["components"]["embeddings"] = f"error: {str(e)}"
        health_status["status"] = "degraded"

    # LLM
    try:
        if rag._llm is not None:
            rag._llm.invoke("ping")
            health_status["components"]["llm"] = "healthy"
        else:
            health_status["components"]["llm"] = "not_loaded"
            health_status["status"] = "degraded"
    except Exception as e:
        health_status["components"]["llm"] = f"error: {str(e)}"
        health_status["status"] = "degraded"

    # Redis cache backend
    try:
        if rag.llm_cache is not None:
            stats = rag.llm_cache.stats()
            health_status["components"]["redis"] = (
                "healthy" if stats.get("backend") == "redis" else "fallback"
            )
        else:
            health_status["components"]["redis"] = "not_loaded"
            health_status["status"] = "degraded"
    except Exception as e:
        health_status["components"]["redis"] = f"error: {str(e)}"
        health_status["status"] = "degraded"

    if all(v == "healthy" for v in health_status["components"].values()):
        health_status["status"] = "healthy"

    return health_status

# ============================================================
# 🔟  ENTRYPOINT
# ============================================================

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)