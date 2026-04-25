# app/ingestion.py
import sys
import os
os.environ["USER_AGENT"] = "langchain-ingestion/1.0"
from typing import List
from langchain_core.documents import Document
from langchain_community.document_loaders import PyPDFLoader
from langchain_community.document_loaders import Docx2txtLoader
import logging
import time

logger = logging.getLogger(__name__)

# Windows: inject Poppler into PATH for this process (pdf2image dependency)
# if os.name == "nt":
#     os.environ["PATH"] = r"C:\Poppler\poppler-25.12.0\Library\bin" + os.pathsep + os.environ.get("PATH", "")

if os.name == "nt":
    # Automatically finds the Library\bin folder relative to your active python.exe
    conda_bin = os.path.join(sys.prefix, 'Library', 'bin')
    os.environ["PATH"] = conda_bin + os.pathsep + os.environ.get("PATH", "")

# ============================================================
# LOADER 1: TEXT FILES
# ============================================================
def load_text_files(directory: str) -> List[Document]:
    from langchain_community.document_loaders import TextLoader

    all_docs = []
    for filename in os.listdir(directory):
        if not filename.endswith(".txt") or filename == "urls.txt":
            continue
        file_path = os.path.join(directory, filename)
        start = time.time()

        try:
            loader = TextLoader(file_path)
            docs = loader.load()
            elapsed = round(time.time() - start, 2)

            total_chars = sum(len(d.page_content) for d in docs)
            logger.info(f"[TXT] ✅ {filename} | {len(docs)} docs | {total_chars} chars | {elapsed}s")

            for doc in docs:
                doc.metadata["source_type"] = "text"
                doc.metadata["file_name"] = filename
            all_docs.extend(docs)

        except Exception as e:
            elapsed = round(time.time() - start, 2)
            logger.error(f"[TXT] ❌ {filename} | Failed after {elapsed}s | {e}")

    return all_docs

# ============================================================
# LOADER 2: PDF FILES (Smart — handles text + scanned)
# ============================================================


def load_pdf(file_path: str) -> List[Document]:
    from langchain_community.document_loaders import PyPDFLoader

    filename = os.path.basename(file_path)
    start = time.time()

    try:
        loader = PyPDFLoader(file_path)
        docs = loader.load()
        elapsed = round(time.time() - start, 2)

        total_chars = sum(len(d.page_content) for d in docs)
        logger.info(f"[PDF] ✅ {filename} | {len(docs)} pages | {total_chars} chars | {elapsed}s")

        for doc in docs:
            doc.metadata["source_type"] = "pdf"
            doc.metadata["file_name"] = filename

        return docs

    except Exception as e:
        elapsed = round(time.time() - start, 2)
        logger.error(f"[PDF] ❌ {filename} | Failed after {elapsed}s | {e}")
        return []

# ============================================================
# LOADER 3: CSV FILES
# ============================================================
def load_csv(file_path: str) -> List[Document]:
    from langchain_community.document_loaders import CSVLoader

    filename = os.path.basename(file_path)
    loader = CSVLoader(file_path)
    docs = loader.load()
    for doc in docs:
        doc.metadata["source_type"] = "csv"
        doc.metadata["file_name"] = filename
    return docs


# ============================================================
# LOADER 4: DOCX FILES
# ============================================================
def load_docx_files(directory: str) -> List[Document]:
    from langchain_community.document_loaders import Docx2txtLoader

    all_docs = []
    for filename in os.listdir(directory):
        if not filename.endswith(".docx"):
            continue
        file_path = os.path.join(directory, filename)
        start = time.time()

        try:
            loader = Docx2txtLoader(file_path)
            docs = loader.load()
            elapsed = round(time.time() - start, 2)

            total_chars = sum(len(d.page_content) for d in docs)
            logger.info(f"[DOCX] ✅ {filename} | {len(docs)} docs | {total_chars} chars | {elapsed}s")

            for doc in docs:
                doc.metadata["source_type"] = "docx"
                doc.metadata["file_name"] = filename
            all_docs.extend(docs)

        except Exception as e:
            elapsed = round(time.time() - start, 2)
            logger.error(f"[DOCX] ❌ {filename} | Failed after {elapsed}s | {e}")

    return all_docs


# ============================================================
# LOADER 5: EXCEL FILES (Custom — multi-sheet, per-row)
# ============================================================
def load_excel_files(directory: str) -> List[Document]:
    import openpyxl

    all_docs = []
    for filename in os.listdir(directory):
        if not filename.endswith(".xlsx"):
            continue
        file_path = os.path.join(directory, filename)
        wb = openpyxl.load_workbook(file_path)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            headers = [cell.value for cell in ws[1]]

            for row in ws.iter_rows(min_row=2, values_only=True):
                # Skip entirely empty rows
                if all(v is None for v in row):
                    continue
                row_data = dict(zip(headers, row))
                page_content = ", ".join(
                    f"{k}: {v}" for k, v in row_data.items() if k is not None
                )
                doc = Document(
                    page_content=page_content,
                    metadata={
                        "source_type": "excel",
                        "file_name": filename,
                        "sheet_name": sheet_name,
                    },
                )
                all_docs.append(doc)

    return all_docs


# ============================================================
# LOADER 6: WEB PAGES
# ============================================================
def load_web_pages(urls: List[str]) -> List[Document]:
    from langchain_community.document_loaders import WebBaseLoader

    loader = WebBaseLoader(urls)
    docs = loader.load()
    for doc in docs:
        doc.metadata["source_type"] = "web"
    return docs


# ============================================================
# MASTER INGESTION FUNCTION
# ============================================================
def ingest_all() -> List[Document]:
    start_total = time.time()

    txt_docs = load_text_files("data/")

    pdf_docs = []
    for filename in os.listdir("data/"):
        if not filename.endswith(".pdf"):
            continue
        file_path = os.path.join("data/", filename)
        pdf_docs.extend(load_pdf(file_path))

    csv_docs = []
    for filename in os.listdir("data/"):
        if not filename.endswith(".csv"):
            continue
        file_path = os.path.join("data/", filename)
        csv_docs.extend(load_csv(file_path))

    docx_docs = load_docx_files("data/")
    excel_docs = load_excel_files("data/")

    # Web pages (optional)
    web_docs = []
    urls_path = "data/urls.txt"
    if os.path.exists(urls_path):
        with open(urls_path, "r") as f:
            urls = [line.strip() for line in f if line.strip()]
        web_docs = load_web_pages(urls)

    all_docs = txt_docs + pdf_docs + csv_docs + docx_docs + excel_docs + web_docs

    elapsed_total = round(time.time() - start_total, 2)

    # ---- Structured Report ----
    logger.info("=" * 60)
    logger.info("📊 INGESTION REPORT")
    logger.info("=" * 60)
    logger.info(f"  📄 Text files:   {len(txt_docs)} documents")
    logger.info(f"  📕 PDF pages:    {len(pdf_docs)} documents")
    logger.info(f"  📊 CSV rows:     {len(csv_docs)} documents")
    logger.info(f"  📝 DOCX files:   {len(docx_docs)} documents")
    logger.info(f"  📗 Excel rows:   {len(excel_docs)} documents")
    logger.info(f"  🌐 Web pages:    {len(web_docs)} documents")
    logger.info("-" * 40)
    logger.info(f"  📦 TOTAL:        {len(all_docs)} documents")
    logger.info(f"  ⏱️  Time:         {elapsed_total}s")
    logger.info("=" * 60)

    return all_docs

# ============================================================
# RUN
# ============================================================
if __name__ == "__main__":
    docs = ingest_all()